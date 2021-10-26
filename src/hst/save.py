import os

from openpyxl import Workbook, load_workbook

from hst.components.subject import SubjectInfo
from hst.data import DataSet, ReportData

HEADER = [
    "DateTime", "Name", "Birth", "Height", "Weight", "Foot", "Lever", "Exercise", "Repetition",
    "Left Max Force", "Left Max Slope", "Left Slope 100ms", "Left Slope 200ms",
    "Right Max Force", "Right Max Slope", "Right Slope 100ms", "Right Slope 200ms",
]


def create_project_file(location: str):
    save_file = os.path.join(location, "measurements.xlsx")
    if not os.path.exists(save_file):
        wb = Workbook()
        sheet = wb.active
        sheet.append(HEADER)
        wb.save(filename=save_file)


def save_run(location: str, subject: SubjectInfo,
             left_report: ReportData, right_report: ReportData):
    save_file = os.path.join(location, "measurements.xlsx")
    new_row = [
        left_report.datetime, subject.name, subject.birth,
        subject.height, subject.weight, subject.foot, subject.lever,
        subject.exercise, subject.repetition,
        left_report.max_force, left_report.max_angel,
        left_report.first_angle, left_report.second_angle,
        right_report.max_force, right_report.max_angel,
        right_report.first_angle, right_report.second_angle,
    ]

    if not os.path.exists(save_file):
        create_project_file(location)

    wb = load_workbook(save_file)
    wb.active.append(new_row)
    wb.save(filename=save_file)


def save_raw_run(location: str, run_datetime: str, subject_name: str,
                 left_raw: DataSet, right_raw: DataSet):
    if not os.path.exists(location):
        return

    save_file = os.path.join(location, f"{run_datetime}_{subject_name}.csv")
    with open(save_file, "w") as f:
        f.write("time,left,right\n")
        for t, l, r in zip(left_raw.x, left_raw.y, right_raw.y):
            f.write(f"{t},{l},{r}\n")
